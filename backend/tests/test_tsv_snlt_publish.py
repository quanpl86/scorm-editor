"""Tests for TSV → SNLT lesson package publish."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.tsv_snlt_publish import (
    TsvPublishError,
    normalize_lesson_code,
    publish_lesson_package,
    split_combined_tsv,
)
from app.quiz_builder import IMPORT_TEMPLATE_DIR

SETTINGS = """Field\tValue\tDescription
title\tQuiz test TSV\tTên
description\tMô tả\t
coverImage\tmedia/quiz_cover.jpg\t
subject\tTin học\t
difficultyLevel\tmedium\t
tags\tSNLT,test\t
createdBy\tagent\t
createdByName\tAI\t
isPublic\tFalse\t
duration\t600\t
shuffleQuestions\tTrue\t
shuffleAnswers\tTrue\t
attemptLimit\t3\t
showResults\tafter_submit\t
allowReview\tTrue\t
createdAt\t\t
updatedAt\t\t
"""

QUESTIONS = """Question Type\tQuestion Text\tAnswer 1\tAnswer 2\tAnswer 3\tAnswer 4\tAnswer 5\tAnswer 6\tExplanation\tDifficulty\tTopic\tPoints\tRequired\tUse Regex\tImage\tVideo\tAudio\tAnswer 1 Image\tAnswer 2 Image\tAnswer 3 Image\tAnswer 4 Image\tAnswer 5 Image\tAnswer 6 Image\tAnswer 1 Left Image\tAnswer 2 Left Image\tAnswer 3 Left Image\tAnswer 4 Left Image\tAnswer 5 Left Image\tAnswer 6 Left Image\tAnswer 1 Right Image\tAnswer 2 Right Image\tAnswer 3 Right Image\tAnswer 4 Right Image\tAnswer 5 Right Image\tAnswer 6 Right Image
MC\tCâu hỏi mẫu?\t*Đúng\tSai\t\t\t\t\tGiải thích\teasy\tLO1|Test\t1\tTrue\tFalse\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t
"""


def test_normalize_lesson_code():
    assert normalize_lesson_code("SNLT-HP01-B02") == "SNLT-HP01-B02"
    with pytest.raises(TsvPublishError):
        normalize_lesson_code("")
    with pytest.raises(TsvPublishError):
        normalize_lesson_code("../evil")


def test_split_combined_tsv():
    blob = f"### quiz_settings.tsv\n{SETTINGS}\n### quiz_questions.tsv\n{QUESTIONS}"
    s, q = split_combined_tsv(blob)
    assert s.startswith("Field\t")
    assert "Question Type" in q


def test_publish_lesson_package(tmp_path: Path):
    root = tmp_path / "ImportTemplate"
    root.mkdir()
    # use real template path but write under tmp
    code = "SNLT-TEST-B99"
    result = publish_lesson_package(
        code,
        SETTINGS,
        QUESTIONS,
        overwrite=False,
        seed_media_from_template=False,
        import_root=root,
        template_xlsx=IMPORT_TEMPLATE_DIR / "SNLT-HP01-B01" / "SNLT-HP01-B01.xlsx",
    )
    assert Path(result["excelPath"]).is_file()
    assert Path(result["mediaDir"]).is_dir()
    assert result["questionCount"] == 1

    wb = load_workbook(result["excelPath"], data_only=True)
    qs = wb["Quiz Questions"]
    assert qs.cell(2, 1).value == "MC"
    assert "mẫu" in str(qs.cell(2, 2).value)
    settings = wb["Quiz Settings"]
    # title field
    titles = {
        str(settings.cell(r, 1).value): settings.cell(r, 2).value
        for r in range(2, settings.max_row + 1)
        if settings.cell(r, 1).value
    }
    assert titles.get("title") == "Quiz test TSV"
    wb.close()

    # Second publish without overwrite fails because xlsx already exists
    with pytest.raises(TsvPublishError, match="đã tồn tại"):
        publish_lesson_package(
            code,
            SETTINGS,
            QUESTIONS,
            overwrite=False,
            import_root=root,
            template_xlsx=IMPORT_TEMPLATE_DIR / "SNLT-HP01-B01" / "SNLT-HP01-B01.xlsx",
        )

    # Empty lesson dir alone does not block
    empty_code = "SNLT-EMPTY-B01"
    (root / empty_code).mkdir()
    r2 = publish_lesson_package(
        empty_code,
        SETTINGS,
        QUESTIONS,
        overwrite=False,
        import_root=root,
        template_xlsx=IMPORT_TEMPLATE_DIR / "SNLT-HP01-B01" / "SNLT-HP01-B01.xlsx",
    )
    assert Path(r2["excelPath"]).is_file()
