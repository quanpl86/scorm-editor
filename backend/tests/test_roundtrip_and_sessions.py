from __future__ import annotations

import json
import io
import os
import threading
import time
import zipfile
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path

from PIL import Image

from app.excel_exporter import export_session_to_excel_zip
from app.excel_import import parse_excel_file, parse_quiz_settings
from app.main import merge_cms_config_into_editor_state, replace_editor_media_with_s3
from app.scorm_parser import atomic_write_text, quiz_to_view
from app.media_bundle import RemoteMediaError, fetch_remote_media
from app import session_manager


def test_merge_cms_configuration_back_into_editor_state():
    cms = {
        "title": "Round trip",
        "description": "Description",
        "subject": "Subject",
        "targetLesson": "Lesson",
        "difficultyLevel": "hard",
        "tags": ["one", "two"],
        "createdBy": "author",
        "createdByName": "Author",
        "isPublic": True,
        "duration": 900,
        "settings": {"shuffleQuestions": True, "attemptLimit": 3},
        "createdAt": "2026-01-01T00:00:00Z",
        "updatedAt": "2026-01-02T00:00:00Z",
        "coverImageUrl": "https://cdn.example.com/cover.png",
    }
    state = {"d": {"T": "Old", "sl": {"g": []}}, "_teky": {"title": "Old"}}

    restored = quiz_to_view(merge_cms_config_into_editor_state(cms, state))["tekyQuiz"]

    for key, value in cms.items():
        if key == "coverImageUrl":
            assert restored["coverImage"] == value
        else:
            assert restored[key] == value


def test_s3_replacement_does_not_duplicate_url_or_filename():
    state = {
        "image": "storage://images/photo.png",
        "html": '<img src="photo.png">',
        "alreadyRemote": "https://s3.example.com/photo.png",
    }
    remote = "https://s3.example.com/photo.png"

    replace_editor_media_with_s3(state, {"photo.png": remote})

    assert state["image"] == remote
    assert state["html"] == f'<img src="{remote}">'
    assert state["alreadyRemote"] == remote


def test_atomic_write_uses_unique_temp_files_under_concurrency(tmp_path):
    target = tmp_path / "quiz_data.json"
    errors: list[Exception] = []

    def write(index: int) -> None:
        try:
            atomic_write_text(target, json.dumps({"writer": index}))
        except Exception as exc:  # pragma: no cover - assertion below reports it
            errors.append(exc)

    with ThreadPoolExecutor(max_workers=20) as executor:
        list(executor.map(write, range(20)))

    assert not errors
    assert isinstance(json.loads(target.read_text())["writer"], int)
    assert not list(tmp_path.glob("*.tmp"))


def test_cleanup_uses_heartbeat_and_two_hour_idle_ttl(tmp_path, monkeypatch):
    monkeypatch.setattr(session_manager, "SESSION_IDLE_TTL_SECONDS", 2 * 3600)
    monkeypatch.setattr(session_manager, "SESSION_STORAGE_LIMIT_BYTES", 1024**3)
    root = tmp_path / "sessions"
    idle_id = "00000000-0000-4000-8000-000000000001"
    active_id = "00000000-0000-4000-8000-000000000002"
    idle = root / idle_id
    active = root / active_id
    idle.mkdir(parents=True)
    active.mkdir(parents=True)
    (idle / "data.bin").write_bytes(b"idle")
    (active / "data.bin").write_bytes(b"active")
    now = time.time()
    os.utime(idle, (now - 3 * 3600, now - 3 * 3600))
    os.utime(active, (now - 3 * 3600, now - 3 * 3600))
    session_manager.touch_session(root, active_id)

    report = session_manager.cleanup_sessions(root, now=now)

    assert report["removed"] == 1
    assert not idle.exists()
    assert active.exists()


def test_cleanup_uses_persisted_activity_not_directory_mtime(tmp_path, monkeypatch):
    monkeypatch.setattr(session_manager, "SESSION_IDLE_TTL_SECONDS", 2 * 3600)
    root = tmp_path / "sessions"
    session_id = "00000000-0000-4000-8000-000000000003"
    session_dir = root / session_id
    session_dir.mkdir(parents=True)
    session_manager.touch_session(root, session_id, event="save")
    activity_path = session_dir / session_manager.ACTIVITY_FILENAME
    activity = json.loads(activity_path.read_text())
    activity["lastAccessedAt"] = time.time() - 3 * 3600
    activity_path.write_text(json.dumps(activity))
    os.utime(session_dir, None)

    report = session_manager.cleanup_sessions(root)

    assert report["removed"] == 1
    assert not session_dir.exists()


def test_cleanup_never_deletes_non_session_marker_files(tmp_path, monkeypatch):
    monkeypatch.setattr(session_manager, "SESSION_IDLE_TTL_SECONDS", 0)
    root = tmp_path / "sessions"
    root.mkdir()
    marker = root / ".gitkeep"
    marker.write_text("keep", encoding="utf-8")

    report = session_manager.cleanup_sessions(root)

    assert report["removed"] == 0
    assert marker.read_text(encoding="utf-8") == "keep"


def test_twenty_concurrent_writers_cannot_overbook_storage(tmp_path, monkeypatch):
    monkeypatch.setattr(session_manager, "SESSION_STORAGE_LIMIT_BYTES", 1000)
    root = tmp_path / "sessions"
    accepted: list[int] = []
    rejected: list[int] = []

    def write(index: int) -> None:
        try:
            with session_manager.storage_write_lock():
                session_manager.ensure_storage_capacity(root, reserve_bytes=100)
                session_dir = root / f"user-{index}"
                session_dir.mkdir(parents=True)
                (session_dir / "data.bin").write_bytes(b"x" * 100)
                session_manager.ensure_storage_capacity(
                    root,
                    protected_ids={session_dir.name},
                )
                accepted.append(index)
        except session_manager.SessionStorageFullError:
            rejected.append(index)

    with ThreadPoolExecutor(max_workers=20) as executor:
        list(executor.map(write, range(20)))

    assert len(accepted) == 10
    assert len(rejected) == 10
    assert session_manager.sessions_size(root) <= 1000


class _ProjectSession:
    def __init__(self, root: Path):
        self.session_id = "project-test"
        self.root = root

    def asset_path(self, value: str) -> Path:
        path = self.root / Path(str(value).split("?")[0]).name
        if not path.exists():
            raise FileNotFoundError(value)
        return path

    def get_view(self):
        image = "question.png"
        second_image = "second.png"
        return {
            "title": "Project Round Trip",
            "questionCount": 5,
            "tekyQuiz": {
                "title": "Project Round Trip",
                "description": "Full settings",
                "coverImage": image,
                "subject": "Subject",
                "targetLesson": "Lesson",
                "lessonCode": "SNLT-TEST",
                "difficultyLevel": "hard",
                "tags": ["roundtrip"],
                "createdBy": "tester",
                "createdByName": "Tester",
                "isPublic": True,
                "duration": 1200,
                "settings": {
                    "shuffleQuestions": True,
                    "shuffleAnswers": True,
                    "attemptLimit": 2,
                    "showResults": "after_submit",
                    "allowReview": True,
                },
            },
            "questions": [
                {
                    "type": "Hotspot",
                    "questionText": "Chọn vùng",
                    "choices": [{
                        "text": "Vùng đúng",
                        "image": image,
                        "isCorrect": True,
                        "rect": {"x": 1000, "y": 1000, "w": 4000, "h": 4000},
                    }],
                    "difficulty": "easy",
                    "points": 1,
                    "slideImages": [image, second_image],
                },
                {
                    "type": "TypeIn",
                    "questionText": "Nhập đáp án",
                    "typeInAnswers": ["A", "B"],
                    "difficulty": "medium",
                    "required": True,
                    "useRegex": True,
                },
                {
                    "type": "FillInTheBlank",
                    "questionText": "___ + ___ = 12",
                    "blankAnswers": [
                        {"id": "blank-1", "values": ["6"]},
                        {"id": "blank-2", "values": ["6", "six"]},
                    ],
                    "wordBankWords": ["4", "8"],
                    "difficulty": "medium",
                },
                {
                    "type": "MultipleNumeric",
                    "questionText": "Hai số",
                    "typeInAnswers": ["2", "3"],
                    "difficulty": "hard",
                },
                {
                    "type": "WordBank",
                    "questionText": "Chọn từ",
                    "blankAnswers": [{"values": ["đúng"]}],
                    "wordBankWords": ["sai"],
                    "difficulty": "easy",
                },
            ],
        }


def test_remote_media_is_downloaded_into_source_project(tmp_path, monkeypatch):
    image_buffer = io.BytesIO()
    Image.new("RGB", (20, 20), "blue").save(image_buffer, format="PNG")
    remote_url = "https://cdn.example.com/question.png"

    class RemoteProjectSession(_ProjectSession):
        def asset_path(self, value: str) -> Path:
            raise FileNotFoundError(value)

        def get_view(self):
            return {
                "title": "Remote Project",
                "tekyQuiz": {"title": "Remote Project", "coverImage": remote_url},
                "questions": [{
                    "id": "remote-q1",
                    "type": "MultipleChoice",
                    "questionText": "Remote?",
                    "slideImages": [remote_url],
                    "choices": [
                        {"text": "A", "isCorrect": True, "image": remote_url},
                        {"text": "B", "isCorrect": False},
                    ],
                    "layout": {"objects": [{
                        "image": remote_url,
                        "r": {"x": 10, "y": 20, "w": 100, "h": 80},
                    }]},
                }],
            }

    monkeypatch.setattr(
        "app.media_bundle.fetch_remote_media",
        lambda _url: (image_buffer.getvalue(), ".png", "image/png"),
    )
    zip_path, _ = export_session_to_excel_zip(RemoteProjectSession(tmp_path))
    try:
        with zipfile.ZipFile(zip_path) as archive:
            media = [name for name in archive.namelist() if name.startswith("media/")]
            assert media
            archive.extractall(tmp_path / "remote-export")
        workbook = next((tmp_path / "remote-export").glob("*.xlsx"))
        rows = parse_excel_file(workbook)
        assert rows[0].image.startswith("media/")
        assert rows[0].extra_media[0]["x"] == 10
        assert rows[0].extra_media[0]["y"] == 20
        import openpyxl

        wb = openpyxl.load_workbook(workbook, data_only=True)
        assert "Question Media" in wb.sheetnames
        assert wb["Question Media"].max_row >= 2
    finally:
        Path(zip_path).unlink(missing_ok=True)


def test_remote_media_rejects_private_network():
    def private_resolver(*_args):
        return [(None, None, None, None, ("127.0.0.1", 80))]

    with __import__("pytest").raises(RemoteMediaError, match="mạng nội bộ"):
        fetch_remote_media("http://localhost/private.png", resolver=private_resolver)


def test_source_project_excel_round_trips_types_settings_and_media(tmp_path):
    Image.new("RGB", (100, 100), "red").save(tmp_path / "question.png")
    Image.new("RGB", (80, 80), "green").save(tmp_path / "second.png")
    session = _ProjectSession(tmp_path)
    zip_path, _ = export_session_to_excel_zip(session)
    try:
        with zipfile.ZipFile(zip_path) as archive:
            archive.extractall(tmp_path / "export")
            assert any(name.startswith("media/") for name in archive.namelist())
        workbook = next((tmp_path / "export").glob("*.xlsx"))
        rows = parse_excel_file(workbook)
        settings = parse_quiz_settings(workbook)

        assert len(rows) == 5
        assert not [row.errors for row in rows if row.errors]
        assert not [row.warnings for row in rows if row.warnings]
        assert settings["title"] == "Project Round Trip"
        assert settings["targetLesson"] == "Lesson"
        assert settings["duration"] == 1200
        assert settings["settings"]["shuffleQuestions"] is True
        assert rows[0].image == "media/Project Round Trip_1_IMG-ND1.png"
        assert rows[0].answers[0].image
        assert len(rows[0].extra_media) >= 2
        assert [blank["values"] for blank in rows[2].blank_answers] == [["6"], ["6", "six"]]
        assert rows[2].distractors == ["4", "8"]
    finally:
        Path(zip_path).unlink(missing_ok=True)
