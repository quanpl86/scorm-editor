"""Publish TSV (settings + questions) into SNLT lesson package under ImportTemplate/.

Creates:
  ImportTemplate/{lesson_code}/
    {lesson_code}.xlsx   # filled from SNLT-HP01-B01 template
    media/               # image folder (empty unless seed_media)
"""

from __future__ import annotations

import csv
import io
import re
import shutil
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .quiz_builder import IMPORT_TEMPLATE_DIR, PROJECT_ROOT

SNLT_TEMPLATE_XLSX = (
    IMPORT_TEMPLATE_DIR / "SNLT-HP01-B01" / "SNLT-HP01-B01.xlsx"
)
SNLT_TEMPLATE_MEDIA = IMPORT_TEMPLATE_DIR / "SNLT-HP01-B01" / "media"

SHEET_QUESTIONS = "Quiz Questions"
SHEET_SETTINGS = "Quiz Settings"

LESSON_CODE_RE = re.compile(r"^[A-Za-z0-9][A-Za-z0-9._-]{1,63}$")

EXPECTED_QUESTION_HEADERS = [
    "Question Type",
    "Question Text",
    "Answer 1",
    "Answer 2",
    "Answer 3",
    "Answer 4",
    "Answer 5",
    "Answer 6",
    "Explanation",
    "Difficulty",
    "Topic",
    "Points",
    "Required",
    "Use Regex",
    "Image",
    "Video",
    "Audio",
    "Answer 1 Image",
    "Answer 2 Image",
    "Answer 3 Image",
    "Answer 4 Image",
    "Answer 5 Image",
    "Answer 6 Image",
    "Answer 1 Left Image",
    "Answer 2 Left Image",
    "Answer 3 Left Image",
    "Answer 4 Left Image",
    "Answer 5 Left Image",
    "Answer 6 Left Image",
    "Answer 1 Right Image",
    "Answer 2 Right Image",
    "Answer 3 Right Image",
    "Answer 4 Right Image",
    "Answer 5 Right Image",
    "Answer 6 Right Image",
]

# Optional schema-v2 columns. Legacy TSV files remain valid; when present these
# preserve multi-holder Fill in Blank mappings and drag-card distractors.
OPTIONAL_QUESTION_HEADERS = ["Blank Answers JSON", "Distractors"]
QUESTION_OUTPUT_HEADERS = EXPECTED_QUESTION_HEADERS + OPTIONAL_QUESTION_HEADERS

EXPECTED_SETTINGS_FIELDS = [
    "title",
    "description",
    "coverImage",
    "subject",  # Related Subject = tên học phần
    "targetLesson",  # Target Lesson = tên bài học
    "difficultyLevel",
    "tags",
    "createdBy",
    "createdByName",
    "isPublic",
    "duration",
    "shuffleQuestions",
    "shuffleAnswers",
    "attemptLimit",
    "showResults",
    "allowReview",
    "createdAt",
    "updatedAt",
]

VALID_TYPES = {"MC", "MR", "TF", "MG", "SEQ", "FIB", "TI", "NUM", "MNUM", "WB"}
VALID_DIFFICULTY = {"easy", "medium", "hard"}
YOUTUBE_RE = re.compile(r"^https?://(www\.)?(youtube\.com/|youtu\.be/)", re.I)
VIMEO_RE = re.compile(r"^https?://(www\.)?vimeo\.com/", re.I)
HTTPS_RE = re.compile(r"^https://", re.I)

# Map alias → easy|medium|hard (key: lower + strip accents)
_DIFFICULTY_ALIASES: dict[str, str] = {
    "easy": "easy",
    "e": "easy",
    "1": "easy",
    "low": "easy",
    "simple": "easy",
    "de": "easy",
    "dễ": "easy",
    "de de": "easy",
    "rat de": "easy",
    "rất dễ": "easy",
    "medium": "medium",
    "med": "medium",
    "m": "medium",
    "2": "medium",
    "normal": "medium",
    "moderate": "medium",
    "mid": "medium",
    "tb": "medium",
    "trung binh": "medium",
    "trung bình": "medium",
    "trungbinh": "medium",
    "vua": "medium",
    "vừa": "medium",
    "hard": "hard",
    "h": "hard",
    "3": "hard",
    "high": "hard",
    "difficult": "hard",
    "kho": "hard",
    "khó": "hard",
    "rat kho": "hard",
    "rất khó": "hard",
}


def _strip_accents(text: str) -> str:
    import unicodedata

    nfkd = unicodedata.normalize("NFKD", text)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def normalize_difficulty(raw: str | None, *, default: str = "medium") -> tuple[str, str | None]:
    """
    Chuẩn hoá Difficulty → easy|medium|hard.
    Returns (normalized, warning_or_None).
    Empty → default medium (warning optional).
    """
    if raw is None:
        return default, None
    s = str(raw).strip()
    if not s:
        return default, None

    low = s.lower()
    if low in VALID_DIFFICULTY:
        return low, None

    # collapse spaces/underscores/hyphens
    key = re.sub(r"[\s_\-]+", " ", low).strip()
    if key in _DIFFICULTY_ALIASES:
        return _DIFFICULTY_ALIASES[key], None

    key_ascii = re.sub(r"[\s_\-]+", " ", _strip_accents(low)).strip()
    if key_ascii in _DIFFICULTY_ALIASES:
        return _DIFFICULTY_ALIASES[key_ascii], None
    # also without spaces
    key_compact = key_ascii.replace(" ", "")
    compact_map = {k.replace(" ", ""): v for k, v in _DIFFICULTY_ALIASES.items()}
    if key_compact in compact_map:
        return compact_map[key_compact], None

    # Unknown: default + warning (do not hard-fail)
    return default, f"Difficulty '{s}' không chuẩn → gán '{default}' (dùng easy|medium|hard)"


class TsvPublishError(ValueError):
    """Validation or I/O error when publishing TSV to lesson package."""


def normalize_lesson_code(raw: str) -> str:
    code = (raw or "").strip()
    if not code:
        raise TsvPublishError("Vui lòng nhập tên Bài học (ví dụ SNLT-HP01-B02).")
    if not LESSON_CODE_RE.match(code):
        raise TsvPublishError(
            "Tên Bài học chỉ gồm chữ, số, dấu chấm, gạch ngang, gạch dưới "
            f"(2–64 ký tự), bắt đầu bằng chữ/số. Nhận: '{code}'"
        )
    if code in {".", ".."} or "/" in code or "\\" in code:
        raise TsvPublishError("Tên Bài học không hợp lệ.")
    return code


def parse_tsv_text(text: str) -> tuple[list[str], list[dict[str, str]]]:
    """Parse TSV string → (fieldnames, rows)."""
    if text is None:
        raise TsvPublishError("TSV rỗng.")
    raw = text.lstrip("\ufeff").replace("\r\n", "\n").replace("\r", "\n")
    lines = [ln for ln in raw.split("\n") if ln.strip() != ""]
    if not lines:
        raise TsvPublishError("TSV rỗng.")

    reader = csv.DictReader(lines, delimiter="\t", restval="")
    if not reader.fieldnames:
        raise TsvPublishError("TSV không có header.")

    fieldnames = [(f or "").strip() for f in reader.fieldnames]
    rows: list[dict[str, str]] = []
    for row in reader:
        clean: dict[str, str] = {}
        for k, v in row.items():
            key = (k or "").strip()
            if not key:
                continue
            clean[key] = "" if v is None else str(v).strip()
        if any(clean.values()):
            rows.append(clean)
    return fieldnames, rows


def split_combined_tsv(blob: str) -> tuple[str, str]:
    """
    Tách một khối paste chứa cả settings + questions.
    Ưu tiên marker:
      ### quiz_settings.tsv
      ### quiz_questions.tsv
    Fallback: nếu có header Field\\tValue → phần settings; phần có Question Type → questions.
    """
    text = (blob or "").lstrip("\ufeff").strip()
    if not text:
        raise TsvPublishError("Chưa dán nội dung TSV.")

    marker_settings = re.compile(
        r"^#{1,6}\s*quiz_settings\.tsv\s*$", re.I | re.M
    )
    marker_questions = re.compile(
        r"^#{1,6}\s*quiz_questions\.tsv\s*$", re.I | re.M
    )

    ms = list(marker_settings.finditer(text))
    mq = list(marker_questions.finditer(text))
    if ms and mq:
        # take first of each
        s_start = ms[0].end()
        q_start = mq[0].end()
        if ms[0].start() < mq[0].start():
            settings = text[s_start : mq[0].start()].strip()
            questions = text[q_start:].strip()
            # trim next marker if any
            next_m = marker_settings.search(questions)
            if next_m:
                questions = questions[: next_m.start()].strip()
        else:
            questions = text[q_start : ms[0].start()].strip()
            settings = text[s_start:].strip()
            next_m = marker_questions.search(settings)
            if next_m:
                settings = settings[: next_m.start()].strip()
        return settings, questions

    # Two fenced blocks ```tsv ... ```
    fences = re.findall(r"```(?:tsv|text)?\s*\n(.*?)```", text, flags=re.I | re.S)
    if len(fences) >= 2:
        a, b = fences[0].strip(), fences[1].strip()
        if a.lower().startswith("field\t") or a.lower().startswith("field,"):
            return a, b
        if b.lower().startswith("field\t"):
            return b, a
        return a, b

    # Heuristic: whole blob is questions only if starts with Question Type
    first = text.split("\n", 1)[0].lower()
    if first.startswith("question type"):
        raise TsvPublishError(
            "Chỉ thấy quiz_questions.tsv. Cần dán thêm quiz_settings.tsv "
            "(hoặc dán 2 khối, đánh dấu ### quiz_settings.tsv / ### quiz_questions.tsv)."
        )
    if first.startswith("field"):
        raise TsvPublishError(
            "Chỉ thấy quiz_settings.tsv. Cần dán thêm quiz_questions.tsv."
        )

    raise TsvPublishError(
        "Không tách được settings/questions. Dán 2 ô riêng, hoặc dùng marker:\n"
        "### quiz_settings.tsv\n...\n### quiz_questions.tsv\n..."
    )


def coerce_cell(header: str, value: str) -> Any:
    if value is None:
        return None
    v = str(value).strip()
    if v == "":
        return None

    if header == "Points":
        try:
            n = float(v)
            return int(n) if n == int(n) else n
        except ValueError:
            return v

    if header in ("duration", "attemptLimit"):
        try:
            return int(float(v))
        except ValueError:
            return v

    if header in ("Difficulty", "difficultyLevel"):
        normalized, _ = normalize_difficulty(v)
        return normalized

    if header in (
        "Required",
        "Use Regex",
        "isPublic",
        "shuffleQuestions",
        "shuffleAnswers",
        "allowReview",
    ):
        low = v.lower()
        if low in ("true", "1", "yes"):
            return True
        if low in ("false", "0", "no"):
            return False
        return v

    return v


_TI_REGEX_GROUP = re.compile(
    r"^\s*\^\(\?i\)\((.+)\)\$\s*$",
    re.I | re.DOTALL,
)
_TI_SIMPLE_PIPE = re.compile(r"^([^|^\n]{1,80}\|)+[^|^\n]{1,80}$")


def _repair_ti_fib_answers(row: dict[str, str], row_no: int) -> str | None:
    """
    Sửa lỗi Agent hay gặp: gộp đồng nghĩa bằng regex/pipe trong Answer 1.

    - Use Regex=True + Answer1 = ^(?i)(a|b|c)$  → giữ regex, xóa Answer2–6
    - Use Regex=False + Answer1 chứa | plain     → tách ra Answer1..N, Use Regex=False
    - Use Regex=False + Answer1 là regex group   → tách alternatives, Use Regex=False
    Returns warning message or None.
    """
    qtype = (row.get("Question Type") or "").strip().upper()
    if qtype not in {"TI", "FIB", "SA"}:
        return None

    a1 = (row.get("Answer 1") or "").strip()
    if not a1:
        return None

    use_rx = (row.get("Use Regex") or "False").strip().lower() in {
        "true", "1", "yes", "y",
    }

    def _clear_extra_answers() -> None:
        for n in range(2, 7):
            row[f"Answer {n}"] = ""

    # Regex mode: keep single-cell regex; clear Answer 2–6 (explanation must not sit here)
    m_rx = _TI_REGEX_GROUP.match(a1)
    if use_rx and m_rx:
        _clear_extra_answers()
        row["Use Regex"] = "True"
        return None

    # Agent put regex but forgot Use Regex=True → prefer expand to synonyms
    if m_rx and not use_rx:
        alts = [p.strip() for p in m_rx.group(1).split("|") if p.strip()]
        alts = alts[:6]
        if len(alts) >= 2:
            for n in range(1, 7):
                row[f"Answer {n}"] = alts[n - 1] if n <= len(alts) else ""
            row["Use Regex"] = "False"
            return (
                f"Hàng {row_no}: TI/FIB Answer 1 là regex nhưng Use Regex=False — "
                f"đã tách thành {len(alts)} đáp án plain text"
            )
        row["Use Regex"] = "True"
        _clear_extra_answers()
        return f"Hàng {row_no}: bật Use Regex=True cho Answer 1 dạng regex"

    # Plain pipe synonyms without regex: BrickColor|Color|Màu sắc
    if (not use_rx) and "|" in a1 and _TI_SIMPLE_PIPE.match(a1) and not a1.startswith("^"):
        alts = [p.strip() for p in a1.split("|") if p.strip()][:6]
        if len(alts) >= 2:
            for n in range(1, 7):
                row[f"Answer {n}"] = alts[n - 1] if n <= len(alts) else ""
            row["Use Regex"] = "False"
            return (
                f"Hàng {row_no}: TI/FIB Answer 1 gộp bằng '|' — "
                f"đã tách thành {len(alts)} cột Answer"
            )

    # Use Regex=True but also filled Answer 2+ — clear extras to protect column layout
    if use_rx:
        extras = any((row.get(f"Answer {n}") or "").strip() for n in range(2, 7))
        if extras:
            _clear_extra_answers()
            return (
                f"Hàng {row_no}: Use Regex=True chỉ dùng Answer 1 — "
                "đã xóa Answer 2–6 để tránh lệch cột"
            )

    return None


def normalize_question_rows(rows: list[dict[str, str]]) -> list[str]:
    """Chuẩn hoá in-place các cột (Difficulty, Type, TI/FIB…) → list warnings."""
    warnings: list[str] = []
    for i, row in enumerate(rows, start=2):
        # Question Type uppercase
        qtype = (row.get("Question Type") or "").strip()
        if qtype:
            row["Question Type"] = qtype.upper()

        raw_diff = row.get("Difficulty")
        if raw_diff is not None and str(raw_diff).strip() != "":
            normalized, warn = normalize_difficulty(raw_diff)
            row["Difficulty"] = normalized
            if warn:
                warnings.append(f"Hàng {i}: {warn}")
        else:
            row["Difficulty"] = "medium"

        # Required / Use Regex normalize True/False strings
        for col in ("Required", "Use Regex"):
            if col not in row:
                continue
            val = (row.get(col) or "").strip()
            if not val:
                row[col] = "False" if col == "Use Regex" else "True"
                continue
            low = val.lower()
            if low in ("true", "1", "yes", "y", "có", "co"):
                row[col] = "True"
            elif low in ("false", "0", "no", "n", "không", "khong"):
                row[col] = "False"

        repair_msg = _repair_ti_fib_answers(row, i)
        if repair_msg:
            warnings.append(repair_msg)
    return warnings


def validate_questions(
    fieldnames: list[str],
    rows: list[dict[str, str]],
    *,
    strict_headers: bool = False,
) -> list[str]:
    warnings: list[str] = []
    errors: list[str] = []
    headers = [h for h in fieldnames if h]
    missing = [h for h in EXPECTED_QUESTION_HEADERS if h not in headers]
    if missing:
        msg = f"Questions TSV thiếu cột: {missing}"
        if strict_headers:
            errors.append(msg)
        else:
            warnings.append(msg)

    if not rows:
        errors.append("Questions TSV không có dòng dữ liệu.")

    # Auto-normalize Difficulty (and related) before hard validation
    warnings.extend(normalize_question_rows(rows))

    for i, row in enumerate(rows, start=2):
        qtype = (row.get("Question Type") or "").strip().upper()
        if not qtype:
            errors.append(f"Hàng {i}: thiếu Question Type")
            continue
        if qtype not in VALID_TYPES:
            errors.append(f"Hàng {i}: Question Type không hợp lệ '{qtype}'")
        if not (row.get("Question Text") or "").strip():
            errors.append(f"Hàng {i}: thiếu Question Text")

        diff = (row.get("Difficulty") or "").strip().lower()
        if diff and diff not in VALID_DIFFICULTY:
            # Should be rare after normalize; still soft-fail to medium
            row["Difficulty"] = "medium"
            warnings.append(
                f"Hàng {i}: Difficulty '{diff}' không nhận diện → gán medium"
            )

        video = (row.get("Video") or "").strip()
        if video and not (YOUTUBE_RE.search(video) or VIMEO_RE.search(video)):
            warnings.append(f"Hàng {i}: Video nên là URL YouTube/Vimeo")

        audio = (row.get("Audio") or "").strip()
        if audio and not HTTPS_RE.search(audio):
            warnings.append(f"Hàng {i}: Audio nên là URL HTTPS trực tiếp")
        if audio.startswith("media/"):
            errors.append(f"Hàng {i}: Audio không dùng file trong media/")

        for col in (
            ["Image"]
            + [f"Answer {n} Image" for n in range(1, 7)]
            + [f"Answer {n} Left Image" for n in range(1, 7)]
            + [f"Answer {n} Right Image" for n in range(1, 7)]
        ):
            val = (row.get(col) or "").strip()
            if not val:
                continue
            if val.startswith("http://") or val.startswith("https://"):
                errors.append(f"Hàng {i} {col}: ảnh dùng path media/…, không URL")
            elif val.lower().endswith((".mp3", ".mp4", ".wav", ".webm", ".mov")):
                errors.append(f"Hàng {i} {col}: media/ chỉ chứa ảnh")

    if errors:
        raise TsvPublishError("Validation Questions thất bại:\n- " + "\n- ".join(errors))
    return warnings


def validate_settings(fieldnames: list[str], rows: list[dict[str, str]]) -> list[str]:
    warnings: list[str] = []
    if "Field" not in fieldnames or "Value" not in fieldnames:
        raise TsvPublishError(
            "quiz_settings.tsv cần cột Field và Value "
            f"(có: {fieldnames})"
        )
    fields = {(r.get("Field") or "").strip() for r in rows if (r.get("Field") or "").strip()}
    if "title" not in fields:
        warnings.append("Settings thiếu field title — sẽ giữ/ghi từ template.")
    missing = [f for f in EXPECTED_SETTINGS_FIELDS if f not in fields]
    if missing:
        warnings.append(f"Settings thiếu field (giữ template): {missing}")
    return warnings


def _clear_sheet_data(ws, start_row: int = 2) -> None:
    if ws.max_row < start_row:
        return
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.value = None


_SETTINGS_FIELD_DESC: dict[str, str] = {
    "title": "Tên hiển thị của quiz",
    "description": "Mô tả quiz",
    "coverImage": "Ảnh đại diện cấp quiz; khác với cột Image của từng câu hỏi",
    "subject": "Related Subject — tên học phần",
    "targetLesson": "Target Lesson — tên bài học",
    "difficultyLevel": "easy | medium | hard",
    "tags": "Các tag phân cách bằng dấu phẩy",
    "createdBy": "Mã người tạo; không phải Quiz ID",
    "createdByName": "Tên người tạo",
    "isPublic": "Quiz có công khai hay không",
    "duration": "Thời lượng làm bài, đơn vị giây",
    "shuffleQuestions": "Trộn thứ tự câu hỏi",
    "shuffleAnswers": "Trộn thứ tự đáp án",
    "attemptLimit": "Số lần làm bài tối đa; 0 = không giới hạn",
    "showResults": "after_submit | immediately | never",
    "allowReview": "Cho phép xem lại sau khi nộp",
    "createdAt": "ISO-8601; để trống để hệ thống tự sinh",
    "updatedAt": "ISO-8601; để trống để hệ thống tự sinh",
}


def apply_settings_to_sheet(ws, rows: list[dict[str, str]]) -> None:
    """
    Ghi đè toàn bộ sheet Quiz Settings theo TSV (không merge giữ data mẫu template).

    Trước đây chỉ update field có trong TSV → các field còn lại giữ nguyên
    giá trị mẫu SNLT-HP01-B01 (subject/tags/duration…), khiến Excel trông
    như «không nhận» cấu hình form/TSV.
    """
    incoming: dict[str, tuple[Any, str | None]] = {}
    for r in rows:
        field = (r.get("Field") or "").strip()
        if not field or field.lower() == "field":
            continue
        val = r.get("Value")
        if val is None:
            val = ""
        else:
            val = str(val)
        desc = r.get("Description")
        if field in ("createdAt", "updatedAt") and not str(val).strip():
            coerced = None
        else:
            coerced = coerce_cell(field, val)
        desc_out = (
            str(desc).strip()
            if desc is not None and str(desc).strip() != ""
            else _SETTINGS_FIELD_DESC.get(field)
        )
        incoming[field] = (coerced, desc_out)

    # Xóa hết hàng data cũ (giữ header)
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)

    ws.cell(1, 1).value = "Field"
    ws.cell(1, 2).value = "Value"
    ws.cell(1, 3).value = "Description"

    # Ghi theo thứ tự chuẩn + field lạ (nếu có)
    ordered_fields = list(EXPECTED_SETTINGS_FIELDS)
    for field in incoming:
        if field not in ordered_fields:
            ordered_fields.append(field)

    for i, field in enumerate(ordered_fields, start=2):
        val, desc = incoming.get(field, (None, _SETTINGS_FIELD_DESC.get(field)))
        ws.cell(i, 1).value = field
        ws.cell(i, 2).value = val
        ws.cell(i, 3).value = desc


def apply_questions_to_sheet(ws, rows: list[dict[str, str]]) -> int:
    col_index: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(1, c).value
        if h is None:
            continue
        col_index[str(h).strip()] = c

    # Đảm bảo đủ header chuẩn (template thiếu cột → thêm)
    for header in QUESTION_OUTPUT_HEADERS:
        if header not in col_index:
            col = (ws.max_column or 0) + 1
            ws.cell(1, col).value = header
            col_index[header] = col

    # Xóa sạch data cũ (kể cả hàng mẫu template) — tránh sót câu hỏi/sample cũ
    _clear_sheet_data(ws, start_row=2)

    for i, row in enumerate(rows):
        excel_row = 2 + i
        for header in QUESTION_OUTPUT_HEADERS:
            col = col_index[header]
            raw = row.get(header, "")
            # DictReader may miss keys if TSV short; also try case-insensitive
            if raw == "" or raw is None:
                for k, v in row.items():
                    if k and k.strip().lower() == header.lower():
                        raw = v
                        break
            ws.cell(excel_row, col).value = coerce_cell(
                header, raw if raw is not None else ""
            )
    return len(rows)


def fill_workbook_from_tsv(
    template_path: Path,
    output_path: Path,
    settings_tsv: str,
    questions_tsv: str,
    *,
    strict_headers: bool = False,
) -> dict[str, Any]:
    """Copy template → output and fill Settings + Questions from TSV text."""
    if not template_path.is_file():
        raise TsvPublishError(f"Không tìm thấy template: {template_path}")

    s_fields, s_rows = parse_tsv_text(settings_tsv)
    q_fields, q_rows = parse_tsv_text(questions_tsv)
    warnings = validate_settings(s_fields, s_rows)
    warnings.extend(
        validate_questions(q_fields, q_rows, strict_headers=strict_headers)
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template_path, output_path)

    wb = load_workbook(output_path)
    try:
        if SHEET_SETTINGS not in wb.sheetnames:
            raise TsvPublishError(f"Template thiếu sheet {SHEET_SETTINGS}")
        if SHEET_QUESTIONS not in wb.sheetnames:
            raise TsvPublishError(f"Template thiếu sheet {SHEET_QUESTIONS}")
        apply_settings_to_sheet(wb[SHEET_SETTINGS], s_rows)
        n = apply_questions_to_sheet(wb[SHEET_QUESTIONS], q_rows)
        wb.save(output_path)
    finally:
        wb.close()

    return {
        "questionCount": n,
        "settingsFields": len(s_rows),
        "warnings": warnings,
        "excelPath": str(output_path),
    }


_MEDIA_PATH_RE = re.compile(
    r"(?:^|[\s=\"'\[\]|,;])(media/[A-Za-z0-9_./\-]+\.(?:jpg|jpeg|png|gif|webp|bmp|mp3|wav|m4a|ogg|mp4|webm|mov))",
    re.I,
)


def collect_media_refs(settings_tsv: str, questions_tsv: str) -> set[str]:
    """Thu thập path media/... được khai báo trong settings + questions TSV."""
    refs: set[str] = set()
    blob = f"{settings_tsv or ''}\n{questions_tsv or ''}"
    for m in _MEDIA_PATH_RE.finditer(blob):
        refs.add(m.group(1).replace("\\", "/"))
    # cover mặc định form
    if "coverImage" in blob and "media/" not in blob:
        refs.add("media/quiz_cover.jpg")
    # luôn thử có cover nếu settings có coverImage rỗng sau parse — add default
    try:
        _, s_rows = parse_tsv_text(settings_tsv)
        for r in s_rows:
            if (r.get("Field") or "").strip() == "coverImage":
                val = (r.get("Value") or "").strip()
                if val.startswith("media/"):
                    refs.add(val)
                elif not val:
                    refs.add("media/quiz_cover.jpg")
                break
        else:
            refs.add("media/quiz_cover.jpg")
    except TsvPublishError:
        refs.add("media/quiz_cover.jpg")
    return refs


def ensure_lesson_media(
    media_dir: Path,
    *,
    media_src: Path,
    settings_tsv: str,
    questions_tsv: str,
    seed_all: bool = False,
) -> dict[str, Any]:
    """
    Đảm bảo thư mục media/ có file cần thiết.
    - seed_all=True: copy toàn bộ media template
    - mặc định: copy từng file media/* được tham chiếu (và quiz_cover.jpg) từ template seed
    """
    media_dir.mkdir(parents=True, exist_ok=True)
    copied: list[str] = []
    missing: list[str] = []

    if seed_all and media_src.is_dir():
        for src in media_src.iterdir():
            if not src.is_file() or src.name.startswith("."):
                continue
            dest = media_dir / src.name
            if not dest.exists() or seed_all:
                shutil.copy2(src, dest)
                copied.append(src.name)
        return {"copied": copied, "missing": missing, "seedAll": True}

    refs = collect_media_refs(settings_tsv, questions_tsv)
    # luôn đảm bảo cover mặc định
    refs.add("media/quiz_cover.jpg")

    for ref in sorted(refs):
        name = Path(ref).name
        dest = media_dir / name
        if dest.is_file():
            continue
        src = media_src / name if media_src.is_dir() else None
        if src and src.is_file():
            shutil.copy2(src, dest)
            copied.append(name)
        else:
            missing.append(ref)

    # remove empty gitkeep if we have real files
    gitkeep = media_dir / ".gitkeep"
    if gitkeep.exists() and any(p.is_file() and p.name != ".gitkeep" for p in media_dir.iterdir()):
        try:
            gitkeep.unlink()
        except OSError:
            pass

    return {"copied": copied, "missing": missing, "seedAll": False}


def publish_lesson_package(
    lesson_code: str,
    settings_tsv: str,
    questions_tsv: str,
    *,
    overwrite: bool = False,
    seed_media_from_template: bool = False,
    template_xlsx: Path | None = None,
    template_media: Path | None = None,
    import_root: Path | None = None,
) -> dict[str, Any]:
    """
    Tạo gói bài học:
      {import_root}/{lesson_code}/{lesson_code}.xlsx
      {import_root}/{lesson_code}/media/
    """
    code = normalize_lesson_code(lesson_code)
    root = import_root or IMPORT_TEMPLATE_DIR
    lesson_dir = root / code
    excel_path = lesson_dir / f"{code}.xlsx"
    media_dir = lesson_dir / "media"
    template = template_xlsx or SNLT_TEMPLATE_XLSX
    media_src = template_media or SNLT_TEMPLATE_MEDIA

    # Chỉ chặn khi đã có file Excel; thư mục rỗng / chỉ media chưa coi là "đã publish"
    if excel_path.is_file() and not overwrite:
        raise TsvPublishError(
            f"File Excel đã tồn tại: ImportTemplate/{code}/{code}.xlsx. "
            "Bật «Ghi đè nếu thư mục bài học đã tồn tại» trên form, "
            "hoặc chọn tên Bài học khác."
        )

    lesson_dir.mkdir(parents=True, exist_ok=True)

    result = fill_workbook_from_tsv(
        template,
        excel_path,
        settings_tsv,
        questions_tsv,
    )

    media_info = ensure_lesson_media(
        media_dir,
        media_src=media_src,
        settings_tsv=settings_tsv,
        questions_tsv=questions_tsv,
        seed_all=seed_media_from_template,
    )
    if media_info.get("copied"):
        result.setdefault("warnings", []).append(
            f"Đã copy media từ template: {', '.join(media_info['copied'][:12])}"
            + ("…" if len(media_info["copied"]) > 12 else "")
        )
    if media_info.get("missing"):
        result.setdefault("warnings", []).append(
            "Thiếu file media (chưa có trong template seed): "
            + ", ".join(media_info["missing"][:12])
        )

    result.update({
        "lessonCode": code,
        "lessonDir": str(lesson_dir),
        "excelPath": str(excel_path),
        "mediaDir": str(media_dir),
        "relativeExcel": f"ImportTemplate/{code}/{code}.xlsx",
        "relativeMedia": f"ImportTemplate/{code}/media",
        "projectRoot": str(PROJECT_ROOT),
        "overwrite": overwrite,
        "seedMedia": seed_media_from_template,
        "mediaCopied": media_info.get("copied") or [],
        "mediaMissing": media_info.get("missing") or [],
    })
    return result


def tsv_to_string_from_rows(fieldnames: list[str], rows: list[dict[str, str]]) -> str:
    """Helper for tests / round-trip."""
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=fieldnames, delimiter="\t", lineterminator="\n", extrasaction="ignore")
    w.writeheader()
    for row in rows:
        w.writerow({k: row.get(k, "") for k in fieldnames})
    return buf.getvalue()
