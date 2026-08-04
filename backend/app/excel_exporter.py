"""Export an editor session as the canonical Teky Excel + media project ZIP."""

from __future__ import annotations

import io
import json
import os
import tempfile
import zipfile
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from .media_bundle import MediaBundler
from .fill_blank import align_blank_answers, normalize_blank_answers
from .scorm_parser import ScormSession


MAX_ANSWERS = 6
IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".gif", ".bmp", ".webp"}

QUESTION_HEADERS = [
    "Question Type",
    "Question Text",
    *[f"Answer {i}" for i in range(1, MAX_ANSWERS + 1)],
    "Explanation",
    "Difficulty",
    "Topic",
    "Points",
    "Required",
    "Use Regex",
    "Blank Answers JSON",
    "Distractors",
    "Image",
    "Video",
    "Audio",
    *[f"Answer {i} Image" for i in range(1, MAX_ANSWERS + 1)],
    *[f"Answer {i} Left Image" for i in range(1, MAX_ANSWERS + 1)],
    *[f"Answer {i} Right Image" for i in range(1, MAX_ANSWERS + 1)],
]

TYPE_MAP = {
    "MultipleChoice": "MC",
    "MultipleChoiceText": "MC",
    "MultipleResponse": "MR",
    "TrueFalse": "TF",
    "TypeIn": "TI",
    "Sequence": "SEQ",
    "Matching": "MG",
    "FillInTheBlank": "FIB",
    "WordBank": "WB",
    "InfoSlide": "IS",
    "Numeric": "NUM",
    "MultipleNumeric": "MNUM",
    # Teky Excel has no native drag/drop representation. Preserve the visible
    # prompt/media as an information slide instead of producing a skipped row.
    "DND": "IS",
}


def _safe_filename(value: str, fallback: str = "Quiz") -> str:
    cleaned = "".join(c if c.isalnum() or c in " _-" else "_" for c in (value or "").strip())
    return cleaned[:120].strip() or fallback


def _answer_values(q: dict[str, Any], excel_type: str) -> list[dict[str, Any]]:
    """Normalize every supported editor type to at most six canonical Excel answers."""
    if q.get("type") == "Hotspot":
        return [
            {
                "text": choice.get("text", f"Vùng {idx + 1}"),
                "correct": bool(choice.get("isCorrect")),
                "image": choice.get("image"),
                "hotspotRect": choice.get("rect") or {},
            }
            for idx, choice in enumerate(q.get("choices") or [])
        ][:MAX_ANSWERS]

    if q.get("type") == "DND":
        items = q.get("dndItems") or []
        mapped = [item for item in items if item.get("isMapped") and item.get("targetId")]
        target_ids = list(dict.fromkeys(str(item.get("targetId")) for item in mapped))
        if len(target_ids) == 1:
            target_id = target_ids[0]
            return [
                {
                    "text": item.get("sourceText", ""),
                    "image": item.get("sourceImage"),
                    "correct": item.get("targetId") == target_id,
                }
                for item in items[:MAX_ANSWERS]
            ]
        return [
            {
                "text": f"{item.get('sourceText', '')} | {item.get('targetText', '')}",
                "leftImage": item.get("sourceImage"),
                "rightImage": item.get("targetImage"),
            }
            for item in mapped[:MAX_ANSWERS]
        ]

    if excel_type == "MG":
        return [
            {
                "text": f"{pair.get('premise', '')} | {pair.get('response', '')}",
                "leftImage": pair.get("leftImage") or pair.get("image"),
                "rightImage": pair.get("rightImage"),
            }
            for pair in (q.get("matchingPairs") or [])[:MAX_ANSWERS]
        ]

    if excel_type == "SEQ":
        items = q.get("sequenceItems") or q.get("choices") or []
        return [
            {"text": item.get("text", ""), "image": item.get("image")}
            for item in items[:MAX_ANSWERS]
        ]

    if excel_type in {"TI", "NUM", "MNUM"}:
        return [{"text": str(value), "correct": True} for value in (q.get("typeInAnswers") or [])[:MAX_ANSWERS]]

    if excel_type == "FIB":
        values: list[str] = []
        for blank in q.get("blankAnswers") or []:
            for value in blank.get("values") or blank.get("acceptedAnswers") or []:
                if value and str(value) not in values:
                    values.append(str(value))
        return [{"text": value, "correct": True} for value in values[:MAX_ANSWERS]]

    if excel_type == "WB":
        correct: list[str] = []
        for blank in q.get("blankAnswers") or []:
            for value in blank.get("values") or blank.get("acceptedAnswers") or []:
                if value and str(value) not in correct:
                    correct.append(str(value))
        extras = [str(value) for value in (q.get("wordBankWords") or []) if value and str(value) not in correct]
        return (
            [{"text": value, "correct": True} for value in correct]
            + [{"text": value, "correct": False} for value in extras]
        )[:MAX_ANSWERS]

    return [
        {
            "text": choice.get("text", ""),
            "correct": bool(choice.get("isCorrect")),
            "image": choice.get("image"),
        }
        for choice in (q.get("choices") or [])[:MAX_ANSWERS]
    ]


def export_session_to_excel_zip(session: ScormSession) -> tuple[Path, str]:
    view = session.get_view()
    teky = view.get("tekyQuiz") or {}
    safe_title = _safe_filename(teky.get("title") or view.get("title") or "Quiz")

    fd, temp_path = tempfile.mkstemp(suffix=".zip")
    os.close(fd)

    wb = openpyxl.Workbook()
    ws_questions = wb.active
    ws_questions.title = "Quiz Questions"
    ws_questions.append(QUESTION_HEADERS)
    for cell in ws_questions[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    warnings: list[str] = []
    question_media_rows: list[list[Any]] = []

    with zipfile.ZipFile(temp_path, "w", zipfile.ZIP_DEFLATED) as zf:
        bundler = MediaBundler(session, zf)

        def add_local_media(original: str | None, proposed_name: str) -> str:
            return bundler.add(original, _safe_filename(proposed_name, "media"))

        def add_hotspot_crop(answer: dict[str, Any], proposed_name: str) -> str:
            image_ref = answer.get("image")
            rect = answer.get("hotspotRect") or {}
            if not image_ref or not rect:
                return add_local_media(image_ref, proposed_name)
            try:
                from PIL import Image

                resolved = bundler.read(str(image_ref))
                if not resolved:
                    raise FileNotFoundError(str(image_ref))
                image_bytes, source_suffix = resolved
                with Image.open(io.BytesIO(image_bytes)) as image:
                    x = max(0, int(float(rect.get("x", 0)) * image.width / 10000))
                    y = max(0, int(float(rect.get("y", 0)) * image.height / 10000))
                    width = max(1, int(float(rect.get("w", 0)) * image.width / 10000))
                    height = max(1, int(float(rect.get("h", 0)) * image.height / 10000))
                    crop = image.crop((x, y, min(image.width, x + width), min(image.height, y + height)))
                    suffix = source_suffix if source_suffix in IMAGE_EXTS else ".png"
                    if crop.mode in {"RGBA", "LA"} and suffix in {".jpg", ".jpeg"}:
                        crop = crop.convert("RGB")
                    archive_name = f"media/{_safe_filename(proposed_name, 'hotspot')}{suffix}"
                    buffer = io.BytesIO()
                    fmt = "JPEG" if suffix in {".jpg", ".jpeg"} else ("GIF" if suffix == ".gif" else "PNG")
                    crop.save(buffer, format=fmt)
                    if archive_name not in bundler.names:
                        zf.writestr(archive_name, buffer.getvalue())
                        bundler.names.add(archive_name)
                    return archive_name
            except Exception as exc:
                warnings.append(f"Không crop được hotspot {proposed_name}: {exc}")
                return add_local_media(image_ref, proposed_name)

        for q_index, q in enumerate(view.get("questions") or [], start=1):
            source_type = q.get("type", "")
            answers = _answer_values(q, TYPE_MAP.get(source_type, source_type))
            if source_type == "Hotspot":
                excel_type = "MC" if sum(bool(answer.get("correct")) for answer in answers) <= 1 else "MR"
                if excel_type == "MC" and len(answers) == 1:
                    answers.append({"text": "Khu vực khác", "correct": False})
            elif source_type == "DND":
                mapped_targets = {
                    str(item.get("targetId"))
                    for item in (q.get("dndItems") or [])
                    if item.get("isMapped") and item.get("targetId")
                }
                if len(mapped_targets) == 1:
                    correct_count = sum(bool(answer.get("correct")) for answer in answers)
                    excel_type = "MC" if correct_count <= 1 else "MR"
                    if len(answers) == 1:
                        answers.append({"text": "Phương án khác", "correct": False})
                elif len(mapped_targets) > 1:
                    excel_type = "MG"
                else:
                    excel_type = "IS"
                    answers = []
            elif source_type == "MultipleChoiceText" and len(answers) < 2:
                # This iSpring-only type is not representable in the canonical
                # Teky schema when the source exposes no choices.
                excel_type = "IS"
                answers = []
            else:
                excel_type = TYPE_MAP.get(source_type, source_type)

            prefix = f"{safe_title}_{q_index}"
            question_media: list[str] = []
            seen_question_refs: set[str] = set()

            def add_question_media(reference: str | None, role: str, position: int, rect: dict[str, Any] | None = None) -> str:
                raw = str(reference or "").strip()
                if not raw:
                    return ""
                code = {"image": "IMG", "video": "VID", "audio": "AUD"}.get(role, "MEDIA")
                exported = add_local_media(raw, f"{prefix}_{code}-ND{position}")
                if exported:
                    r = rect or {}
                    question_media_rows.append([
                        q_index,
                        q.get("id", ""),
                        role,
                        position,
                        exported,
                        r.get("x", ""), r.get("y", ""), r.get("w", ""), r.get("h", ""),
                    ])
                return exported

            # Layout objects go first so the portable metadata retains their
            # exact canvas rectangle when the same asset is also a slide image.
            for obj in (q.get("layout") or {}).get("objects") or []:
                for role in ("image", "video", "audio"):
                    value = obj.get(role)
                    raw = str(value or "")
                    if not raw or raw in seen_question_refs:
                        continue
                    seen_question_refs.add(raw)
                    position = 1 + sum(1 for row in question_media_rows if row[0] == q_index and row[2] == role)
                    exported = add_question_media(value, role, position, obj.get("r"))
                    if role == "image" and exported:
                        question_media.append(exported)
            for image in q.get("slideImages") or []:
                raw = str(image)
                if raw in seen_question_refs:
                    continue
                seen_question_refs.add(raw)
                exported = add_question_media(image, "image", len(question_media) + 1)
                if exported:
                    question_media.append(exported)

            video = q.get("video") or ""
            audio = q.get("audio") or ""
            for obj in (q.get("layout") or {}).get("objects") or []:
                video = video or obj.get("video") or ""
                audio = audio or obj.get("audio") or ""
            video_ref = add_local_media(video, f"{prefix}_VID-ND")
            audio_ref = add_local_media(audio, f"{prefix}_AUD-ND")

            explanation = q.get("explanation") or (q.get("feedback") or {}).get("correct") or ""
            feedback = q.get("feedback") or {}
            for kind, key, code in (
                ("image", "correctImage", "IMG-GT"),
                ("audio", "correctAudio", "AUD-GT"),
                ("video", "correctVideo", "VID-GT"),
            ):
                value = feedback.get(key)
                if value:
                    ref = add_local_media(value, f"{prefix}_{code}")
                    if ref:
                        explanation = f"{explanation} [{kind}={ref}]".strip()

            row: dict[str, Any] = {
                "Question Type": excel_type,
                "Question Text": q.get("questionText", ""),
                "Explanation": explanation,
                "Difficulty": q.get("difficulty") or "medium",
                "Topic": q.get("topic") or "",
                "Points": q.get("points", 1),
                "Required": bool(q.get("required", False)),
                "Use Regex": bool(q.get("useRegex", False)),
                "Blank Answers JSON": (
                    json.dumps(
                        align_blank_answers(
                            q.get("questionText", ""),
                            normalize_blank_answers(q.get("blankAnswers") or []),
                        ),
                        ensure_ascii=False,
                    )
                    if source_type in {"FillInTheBlank", "WordBank"}
                    else ""
                ),
                "Distractors": (
                    json.dumps(
                        q.get("blankDistractors") or q.get("wordBankWords") or [],
                        ensure_ascii=False,
                    )
                    if source_type in {"FillInTheBlank", "WordBank"}
                    else ""
                ),
                "Image": question_media[0] if question_media else "",
                "Video": video_ref,
                "Audio": audio_ref,
            }

            for answer_index, answer in enumerate(answers, start=1):
                text = str(answer.get("text") or "")
                if answer.get("correct") and excel_type in {"MC", "MR", "TF", "WB"}:
                    text = f"*{text}"
                row[f"Answer {answer_index}"] = text
                if source_type == "Hotspot":
                    row[f"Answer {answer_index} Image"] = add_hotspot_crop(
                        answer, f"{prefix}_IMG-DA{answer_index}"
                    )
                else:
                    row[f"Answer {answer_index} Image"] = add_local_media(
                        answer.get("image"), f"{prefix}_IMG-DA{answer_index}"
                    )
                row[f"Answer {answer_index} Left Image"] = add_local_media(
                    answer.get("leftImage"), f"{prefix}_IMG-DA-Left{answer_index}"
                )
                row[f"Answer {answer_index} Right Image"] = add_local_media(
                    answer.get("rightImage"), f"{prefix}_IMG-DA-Right{answer_index}"
                )

            ws_questions.append([row.get(header, "") for header in QUESTION_HEADERS])

        ws_settings = wb.create_sheet("Quiz Settings")
        ws_settings.append(["Field", "Value", "Description"])
        configured_settings = teky.get("settings") or {}
        settings_rows = [
            ("title", teky.get("title") or view.get("title") or "", "Tên bài kiểm tra"),
            ("description", teky.get("description") or "", "Mô tả"),
            ("coverImage", add_local_media(teky.get("coverImage"), f"{safe_title}_IMG-COVER"), "Ảnh bìa"),
            ("subject", teky.get("subject") or "", "Tên học phần"),
            ("targetLesson", teky.get("targetLesson") or "", "Tên bài học"),
            ("lessonCode", teky.get("lessonCode") or "", "Mã package"),
            ("difficultyLevel", teky.get("difficultyLevel") or "medium", "Độ khó"),
            ("tags", ", ".join(teky.get("tags") or []) if isinstance(teky.get("tags"), list) else teky.get("tags") or "", "Tags"),
            ("createdBy", teky.get("createdBy") or "", "Người tạo"),
            ("createdByName", teky.get("createdByName") or "", "Tên người tạo"),
            ("isPublic", bool(teky.get("isPublic", False)), "Công khai"),
            ("duration", int(teky.get("duration", 1800) or 0), "Thời lượng giây"),
            ("shuffleQuestions", bool(configured_settings.get("shuffleQuestions", False)), "Trộn câu"),
            ("shuffleAnswers", bool(configured_settings.get("shuffleAnswers", False)), "Trộn đáp án"),
            ("attemptLimit", int(configured_settings.get("attemptLimit", 1) or 0), "Số lần làm"),
            ("showResults", configured_settings.get("showResults", "after_submit"), "Hiện kết quả"),
            ("allowReview", bool(configured_settings.get("allowReview", True)), "Cho phép xem lại"),
            ("createdAt", teky.get("createdAt") or "", "Thời gian tạo"),
            ("updatedAt", teky.get("updatedAt") or "", "Thời gian cập nhật"),
        ]
        for setting in settings_rows:
            ws_settings.append(setting)

        ws_media = wb.create_sheet("Question Media")
        ws_media.append(["Question Index", "Question ID", "Role", "Position", "Path", "X", "Y", "W", "H"])
        for row in question_media_rows:
            ws_media.append(row)

        warnings.extend(bundler.warnings)
        if warnings:
            ws_warnings = wb.create_sheet("Export Warnings")
            ws_warnings.append(["Warning"])
            for warning in warnings:
                ws_warnings.append([warning])

        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        zf.writestr(f"{safe_title}.xlsx", excel_buffer.getvalue())

    return Path(temp_path), safe_title
