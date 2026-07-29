#!/usr/bin/env python3
"""
tsv_to_snlt_xlsx.py — Đổ quiz_settings.tsv + quiz_questions.tsv vào bản sao
template SNLT-HP01-B01.xlsx (Teky LMS schema v2).

Usage:
  python3 scripts/tsv_to_snlt_xlsx.py \\
    --settings path/to/quiz_settings.tsv \\
    --questions path/to/quiz_questions.tsv \\
    --output path/to/SNLT-HPxx-Byy.xlsx

  # Mặc định lấy template ImportTemplate/SNLT-HP01-B01/SNLT-HP01-B01.xlsx
  # Có thể copy luôn thư mục media/ mẫu:
  python3 scripts/tsv_to_snlt_xlsx.py -s settings.tsv -q questions.tsv -o out.xlsx --copy-media

Dependencies: openpyxl, (stdlib only otherwise)
"""

from __future__ import annotations

import argparse
import csv
import re
import shutil
import sys
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover
    print("Cần cài openpyxl: pip install openpyxl", file=sys.stderr)
    raise SystemExit(1) from exc

# --- Paths mặc định (relative to scorm-editor/ or repo root) ---
SCRIPT_DIR = Path(__file__).resolve().parent
SCORM_EDITOR = SCRIPT_DIR.parent
REPO_ROOT = SCORM_EDITOR.parent

DEFAULT_TEMPLATE = (
    REPO_ROOT / "ImportTemplate" / "SNLT-HP01-B01" / "SNLT-HP01-B01.xlsx"
)
DEFAULT_MEDIA_DIR = REPO_ROOT / "ImportTemplate" / "SNLT-HP01-B01" / "media"

SHEET_QUESTIONS = "Quiz Questions"
SHEET_SETTINGS = "Quiz Settings"

# Header chuẩn 35 cột (schema v2) — thứ tự bắt buộc khi ghi
EXPECTED_QUESTION_HEADERS = [
    "Question Type",
    "Question Text",
    "Answer 1",
    "Answer 2",
    "Answer 3",
    "Answer 4",
    "Answer 5",
    "Answer 6",
    "Explanation",
    "Difficulty",
    "Topic",
    "Points",
    "Required",
    "Use Regex",
    "Image",
    "Video",
    "Audio",
    "Answer 1 Image",
    "Answer 2 Image",
    "Answer 3 Image",
    "Answer 4 Image",
    "Answer 5 Image",
    "Answer 6 Image",
    "Answer 1 Left Image",
    "Answer 2 Left Image",
    "Answer 3 Left Image",
    "Answer 4 Left Image",
    "Answer 5 Left Image",
    "Answer 6 Left Image",
    "Answer 1 Right Image",
    "Answer 2 Right Image",
    "Answer 3 Right Image",
    "Answer 4 Right Image",
    "Answer 5 Right Image",
    "Answer 6 Right Image",
]

EXPECTED_SETTINGS_FIELDS = [
    "title",
    "description",
    "coverImage",
    "subject",
    "targetLesson",
    "difficultyLevel",
    "tags",
    "createdBy",
    "createdByName",
    "isPublic",
    "duration",
    "shuffleQuestions",
    "shuffleAnswers",
    "attemptLimit",
    "showResults",
    "allowReview",
    "createdAt",
    "updatedAt",
]

VALID_TYPES = {"MC", "MR", "TF", "MG", "SEQ", "FIB", "TI", "NUM", "MNUM", "WB"}
VALID_DIFFICULTY = {"easy", "medium", "hard"}
YOUTUBE_RE = re.compile(
    r"^https?://(www\.)?(youtube\.com/|youtu\.be/)", re.I
)
VIMEO_RE = re.compile(r"^https?://(www\.)?vimeo\.com/", re.I)
HTTPS_RE = re.compile(r"^https://", re.I)


def die(msg: str, code: int = 1) -> None:
    print(f"ERROR: {msg}", file=sys.stderr)
    raise SystemExit(code)


def read_tsv_rows(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    """Đọc TSV UTF-8 (có/không BOM) → (fieldnames, rows)."""
    if not path.is_file():
        die(f"Không tìm thấy file TSV: {path}")
    raw = path.read_bytes()
    if raw.startswith(b"\xef\xbb\xbf"):
        raw = raw[3:]
    text = raw.decode("utf-8").replace("\r\n", "\n").replace("\r", "\n")
    lines = [ln for ln in text.split("\n") if ln.strip() != ""]
    if not lines:
        die(f"TSV rỗng: {path}")

    reader = csv.DictReader(lines, delimiter="\t", restval="")
    if not reader.fieldnames:
        die(f"TSV không có header: {path}")

    fieldnames = [(f or "").strip() for f in reader.fieldnames]
    rows: list[dict[str, str]] = []
    for row in reader:
        clean: dict[str, str] = {}
        for k, v in row.items():
            key = (k or "").strip()
            if not key:
                continue
            clean[key] = "" if v is None else str(v).strip()
        if any(clean.values()):
            rows.append(clean)
    return fieldnames, rows


def coerce_cell(header: str, value: str) -> Any:
    """Ép kiểu nhẹ cho openpyxl (Points int, bool-like, còn lại string)."""
    if value is None:
        return None
    v = str(value).strip()
    if v == "":
        return None

    if header == "Points":
        try:
            # allow "1.0"
            n = float(v)
            return int(n) if n == int(n) else n
        except ValueError:
            return v

    if header == "duration" or header == "attemptLimit":
        try:
            return int(float(v))
        except ValueError:
            return v

    if header in ("Required", "Use Regex", "isPublic", "shuffleQuestions", "shuffleAnswers", "allowReview"):
        low = v.lower()
        if low in ("true", "1", "yes"):
            return True
        if low in ("false", "0", "no"):
            return False
        return v

    return v


def validate_questions_headers(headers: list[str], strict: bool) -> list[str]:
    """Trả list warning; die nếu thiếu cột bắt buộc khi strict."""
    warnings: list[str] = []
    # Ignore empty trailing headers
    headers = [h for h in headers if h]
    missing = [h for h in EXPECTED_QUESTION_HEADERS if h not in headers]
    extra = [h for h in headers if h not in EXPECTED_QUESTION_HEADERS]
    if missing:
        msg = f"Questions TSV thiếu cột: {missing}"
        if strict:
            die(msg)
        warnings.append(msg)
    if extra:
        warnings.append(f"Questions TSV có cột thừa (sẽ bỏ qua): {extra}")
    # Order warning
    present = [h for h in EXPECTED_QUESTION_HEADERS if h in headers]
    actual_order = [h for h in headers if h in EXPECTED_QUESTION_HEADERS]
    if present != actual_order:
        # Compare to expected order of those present
        expected_order = [h for h in EXPECTED_QUESTION_HEADERS if h in headers]
        if actual_order != expected_order:
            warnings.append(
                "Thứ tự cột Questions khác template (vẫn map theo tên header)."
            )
    return warnings


def validate_question_rows(rows: list[dict[str, str]], strict_media: bool) -> list[str]:
    warnings: list[str] = []
    errors: list[str] = []

    for i, row in enumerate(rows, start=2):  # Excel-like row (header=1)
        qtype = (row.get("Question Type") or "").strip().upper()
        if not qtype:
            errors.append(f"Hàng {i}: thiếu Question Type")
            continue
        if qtype not in VALID_TYPES:
            errors.append(f"Hàng {i}: Question Type không hợp lệ '{qtype}'")

        if not (row.get("Question Text") or "").strip():
            errors.append(f"Hàng {i}: thiếu Question Text")

        diff = (row.get("Difficulty") or "").strip().lower()
        if diff and diff not in VALID_DIFFICULTY:
            errors.append(f"Hàng {i}: Difficulty phải easy|medium|hard (got '{diff}')")

        pts = row.get("Points") or ""
        if pts != "":
            try:
                p = float(pts)
                if p < 1:
                    warnings.append(f"Hàng {i}: Points < 1 ({pts})")
            except ValueError:
                errors.append(f"Hàng {i}: Points không phải số '{pts}'")

        # Max 6 answers — columns only go to 6; warn if *count for MC/MR
        answers = [row.get(f"Answer {n}") or "" for n in range(1, 7)]
        starred = [a for a in answers if a.strip().startswith("*")]
        if qtype == "MC" and len(starred) != 1:
            warnings.append(f"Hàng {i} MC: nên có đúng 1 đáp án '*' (có {len(starred)})")
        if qtype == "MR" and len(starred) < 2:
            warnings.append(f"Hàng {i} MR: nên có ≥2 đáp án '*' (có {len(starred)})")
        if qtype in ("FIB",) and "___" not in (row.get("Question Text") or ""):
            warnings.append(f"Hàng {i} FIB: Question Text nên có '___'")

        use_rx = (row.get("Use Regex") or "False").strip().lower()
        if use_rx in ("true", "1", "yes") and qtype not in ("FIB", "TI"):
            warnings.append(f"Hàng {i}: Use Regex=True chỉ khuyến nghị cho FIB/TI")

        # Media
        for col in ["Image"] + [f"Answer {n} Image" for n in range(1, 7)] + [
            f"Answer {n} Left Image" for n in range(1, 7)
        ] + [f"Answer {n} Right Image" for n in range(1, 7)]:
            val = (row.get(col) or "").strip()
            if not val:
                continue
            if val.startswith("http://") or val.startswith("https://"):
                errors.append(f"Hàng {i} {col}: ảnh phải là path media/…, không dùng URL")
            elif not val.startswith("media/"):
                warnings.append(f"Hàng {i} {col}: nên bắt đầu bằng media/ (got '{val}')")
            else:
                lower = val.lower()
                if lower.endswith((".mp3", ".mp4", ".wav", ".webm", ".mov", ".m4a")):
                    errors.append(
                        f"Hàng {i} {col}: media/ chỉ chứa ảnh — không dùng '{val}'"
                    )

        video = (row.get("Video") or "").strip()
        if video:
            if not (YOUTUBE_RE.search(video) or VIMEO_RE.search(video)):
                msg = f"Hàng {i} Video: chỉ YouTube/Vimeo URL (got '{video[:60]}…')"
                if strict_media:
                    errors.append(msg)
                else:
                    warnings.append(msg)

        audio = (row.get("Audio") or "").strip()
        if audio:
            if not HTTPS_RE.search(audio):
                msg = f"Hàng {i} Audio: phải là HTTPS trực tiếp (got '{audio[:60]}')"
                if strict_media:
                    errors.append(msg)
                else:
                    warnings.append(msg)
            if audio.startswith("media/"):
                errors.append(f"Hàng {i} Audio: không dùng file trong media/ — dùng URL HTTPS")

    if errors:
        for e in errors:
            print(f"ERROR: {e}", file=sys.stderr)
        die(f"{len(errors)} lỗi validation Questions — dừng ghi file.")

    return warnings


def clear_sheet_data(ws, start_row: int = 2) -> None:
    """Xóa toàn bộ giá trị từ start_row trở đi (giữ style hàng mẫu nếu có thì vẫn clear value)."""
    if ws.max_row < start_row:
        return
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.value = None


def apply_settings(ws, rows: list[dict[str, str]], fieldnames: list[str]) -> list[str]:
    """
    Cập nhật sheet Quiz Settings.
    Hỗ trợ TSV dạng Field/Value/Description (khuyến nghị) hoặc chỉ Field/Value.
    """
    warnings: list[str] = []
    # Build map Field -> (Value, Description?)
    incoming: dict[str, tuple[str, str | None]] = {}
    # Detect format
    if "Field" in fieldnames and "Value" in fieldnames:
        for r in rows:
            field = (r.get("Field") or "").strip()
            if not field:
                continue
            val = r.get("Value") or ""
            desc = r.get("Description")
            incoming[field] = (val, desc)
    else:
        # wide format? not supported — die
        die(
            "quiz_settings.tsv phải có cột Field và Value "
            f"(có: {fieldnames})"
        )

    # Existing rows: col A = Field
    field_to_row: dict[str, int] = {}
    for r in range(2, ws.max_row + 1):
        f = ws.cell(r, 1).value
        if f is not None and str(f).strip():
            field_to_row[str(f).strip()] = r

    for field, (val, desc) in incoming.items():
        coerced = coerce_cell(field, val)
        # createdAt/updatedAt intentionally empty
        if field in ("createdAt", "updatedAt") and (val is None or str(val).strip() == ""):
            coerced = None
        if field in field_to_row:
            rr = field_to_row[field]
            ws.cell(rr, 2).value = coerced
            if desc is not None and str(desc).strip() != "":
                ws.cell(rr, 3).value = str(desc).strip()
        else:
            # append new field row
            rr = ws.max_row + 1
            ws.cell(rr, 1).value = field
            ws.cell(rr, 2).value = coerced
            if desc is not None:
                ws.cell(rr, 3).value = str(desc).strip() if str(desc).strip() else None
            field_to_row[field] = rr
            warnings.append(f"Settings: thêm field mới không có trong template: {field}")

    missing = [f for f in EXPECTED_SETTINGS_FIELDS if f not in incoming]
    if missing:
        warnings.append(
            f"Settings TSV không có các field (giữ nguyên template): {missing}"
        )

    return warnings


def apply_questions(ws, rows: list[dict[str, str]]) -> int:
    """
    Ghi Questions: map theo tên header hàng 1 của workbook (không phụ thuộc thứ tự TSV).
    Xóa data cũ từ hàng 2.
    """
    # Read excel headers
    excel_headers: list[str] = []
    col_index: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(1, c).value
        if h is None:
            continue
        name = str(h).strip()
        excel_headers.append(name)
        col_index[name] = c

    missing = [h for h in EXPECTED_QUESTION_HEADERS if h not in col_index]
    if missing:
        die(f"Template Excel thiếu cột Questions: {missing}")

    clear_sheet_data(ws, start_row=2)

    for i, row in enumerate(rows):
        excel_row = 2 + i
        for header, col in col_index.items():
            raw = row.get(header, "")
            ws.cell(excel_row, col).value = coerce_cell(header, raw if raw is not None else "")

    return len(rows)


def copy_media_dir(src: Path, dest_parent: Path) -> Path | None:
    """Copy media/ cạnh file xlsx output."""
    if not src.is_dir():
        print(f"WARN: không có thư mục media mẫu: {src}", file=sys.stderr)
        return None
    dest = dest_parent / "media"
    if dest.exists():
        shutil.rmtree(dest)
    shutil.copytree(src, dest)
    return dest


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Đổ TSV (settings + questions) vào copy SNLT-HP01-B01.xlsx",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Ví dụ:
  python3 scripts/tsv_to_snlt_xlsx.py \\
    -s ./out/quiz_settings.tsv \\
    -q ./out/quiz_questions.tsv \\
    -o ./work/SNLT-HP01-B01/SNLT-HP01-B01.xlsx \\
    --copy-media
        """,
    )
    p.add_argument(
        "-s",
        "--settings",
        type=Path,
        required=True,
        help="Đường dẫn quiz_settings.tsv",
    )
    p.add_argument(
        "-q",
        "--questions",
        type=Path,
        required=True,
        help="Đường dẫn quiz_questions.tsv",
    )
    p.add_argument(
        "-t",
        "--template",
        type=Path,
        default=DEFAULT_TEMPLATE,
        help=f"Template xlsx (mặc định: {DEFAULT_TEMPLATE})",
    )
    p.add_argument(
        "-o",
        "--output",
        type=Path,
        required=True,
        help="File xlsx đầu ra (sẽ tạo/ghi đè)",
    )
    p.add_argument(
        "--copy-media",
        action="store_true",
        help="Copy ImportTemplate/.../media/ cạnh file output",
    )
    p.add_argument(
        "--media-src",
        type=Path,
        default=DEFAULT_MEDIA_DIR,
        help="Thư mục media nguồn khi --copy-media",
    )
    p.add_argument(
        "--strict-media",
        action="store_true",
        help="Video/Audio sai chuẩn → lỗi (mặc định chỉ warning)",
    )
    p.add_argument(
        "--strict-headers",
        action="store_true",
        help="Thiếu cột Questions → lỗi",
    )
    p.add_argument(
        "--dry-run",
        action="store_true",
        help="Chỉ validate, không ghi file",
    )
    return p.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)

    template: Path = args.template.expanduser().resolve()
    output: Path = args.output.expanduser().resolve()
    settings_path: Path = args.settings.expanduser().resolve()
    questions_path: Path = args.questions.expanduser().resolve()

    if not template.is_file():
        die(f"Không tìm thấy template: {template}")

    print(f"Template : {template}")
    print(f"Settings : {settings_path}")
    print(f"Questions: {questions_path}")
    print(f"Output   : {output}")

    s_fields, s_rows = read_tsv_rows(settings_path)
    q_fields, q_rows = read_tsv_rows(questions_path)

    print(f"Đọc settings: {len(s_rows)} field rows | questions: {len(q_rows)} câu")

    all_warnings: list[str] = []
    all_warnings.extend(validate_questions_headers(q_fields, strict=args.strict_headers))
    all_warnings.extend(
        validate_question_rows(q_rows, strict_media=args.strict_media)
    )

    if not q_rows:
        die("Questions TSV không có dòng dữ liệu.")

    for w in all_warnings:
        print(f"WARN: {w}")

    if args.dry_run:
        print("Dry-run: validation xong, không ghi file.")
        return 0

    output.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template, output)

    wb = load_workbook(output)
    if SHEET_SETTINGS not in wb.sheetnames:
        die(f"Template thiếu sheet '{SHEET_SETTINGS}'")
    if SHEET_QUESTIONS not in wb.sheetnames:
        die(f"Template thiếu sheet '{SHEET_QUESTIONS}'")

    sw = apply_settings(wb[SHEET_SETTINGS], s_rows, s_fields)
    for w in sw:
        print(f"WARN: {w}")

    n = apply_questions(wb[SHEET_QUESTIONS], q_rows)
    wb.save(output)
    wb.close()

    media_dest = None
    if args.copy_media:
        media_dest = copy_media_dir(
            args.media_src.expanduser().resolve(),
            output.parent,
        )

    print("---")
    print(f"OK: đã ghi {n} câu hỏi → {output}")
    if media_dest:
        print(f"OK: media → {media_dest}")
    print(
        "Tiếp theo: gắn/kiểm tra path ảnh media/*, "
        "Video (YouTube/Vimeo), Audio (HTTPS) → zip Excel+media → import Editor."
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
